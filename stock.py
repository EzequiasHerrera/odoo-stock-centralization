import pandas as pd
from openpyxl import load_workbook

def procesar_stock():
    print(">>> Iniciando procesamiento de stock...")

    # Nombres de archivos fijos en carpeta actual
    stock_file = "stock.xlsx"
    tienda_file = "tiendanube.csv"
    eskit_file = "EsKit.xlsx"

    # Leer stock.xlsx
    print("Leyendo archivo de stock:", stock_file)
    stock = pd.read_excel(stock_file)
    stock["default_code"] = stock["default_code"].astype(str).str.strip()
    print(f"Stock cargado con {len(stock)} registros.")

    # Leer tiendanube.csv
    print("Leyendo archivo de Tiendanube:", tienda_file)
    tienda = pd.read_csv(tienda_file, sep=None, engine="python", encoding="latin1", on_bad_lines="skip")
    tienda["SKU"] = tienda["SKU"].astype(str).str.strip()
    print(f"Tiendanube cargado con {len(tienda)} registros.")

    # --- Filtrar registros básicos ---
    print("Aplicando filtros básicos (Comb, Outl, Funsales)...")
    def es_funsales(sku):
        if "|" in str(sku):
            return True
        try:
            int(str(sku))
            return True
        except ValueError:
            return False

    antes = len(stock)
    stock = stock[~stock["default_code"].str.startswith("Comb")]
    stock = stock[~stock["default_code"].str.startswith("Outl")]
    stock = stock[~stock["default_code"].apply(es_funsales)]
    print(f"Filtrado básico: {antes - len(stock)} registros eliminados, {len(stock)} restantes.")

    # --- Cruce por SKU con Tiendanube ---
    print("Actualizando stock desde Tiendanube...")
    tienda_dict = dict(zip(tienda["SKU"], tienda["Stock"]))
    stock["qty_available"] = stock["default_code"].map(tienda_dict).fillna(stock["qty_available"])

    # --- Eliminar registros con stock 0 ---
    antes = len(stock)
    stock = stock[stock["qty_available"].astype(float) != 0]
    print(f"Eliminados {antes - len(stock)} registros con stock 0. Quedan {len(stock)}.")

    # --- Eliminar SKUs marcados como TRUE en EsKit.xlsx ---
    print("Aplicando filtro de kits desde EsKit.xlsx...")
    eskit = pd.read_excel(eskit_file, header=0)
    eskit["Referencia interna"] = eskit["Referencia interna"].astype(str).str.strip()
    eskit["Es un kit"] = eskit["Es un kit"].astype(str).fillna("").str.strip().str.upper()
    eskit_true = set(eskit.loc[eskit["Es un kit"] == "TRUE", "Referencia interna"].tolist())

    stock = stock[~stock["default_code"].isin(eskit_true)]
    print(f"Eliminados {len(eskit_true)} SKUs por ser kits. Quedan {len(stock)} registros finales.")

    # --- Guardar resultado en bloques de 1000 registros ---
    wb = load_workbook(stock_file)
    # Eliminar hojas previas Importar
    for sheet in wb.sheetnames:
        if sheet.startswith("Importar"):
            wb.remove(wb[sheet])

    max_rows = 1000
    total = len(stock)
    num_sheets = (total // max_rows) + (1 if total % max_rows else 0)

    print(f"Generando {num_sheets} hojas de hasta {max_rows} registros cada una...")

    for i in range(num_sheets):
        start = i * max_rows
        end = start + max_rows
        chunk = stock.iloc[start:end]

        sheet_name = f"Importar_{i+1}"
        ws_out = wb.create_sheet(sheet_name)
        ws_out.append(list(stock.columns))

        for j, row in enumerate(chunk.itertuples(index=False), start=1):
            ws_out.append(list(row))
            if j % 100 == 0:
                print(f"   > {j} registros escritos en {sheet_name}...")

    wb.save(stock_file)
    print(">>> Procesamiento terminado.")
    print(f"Resultados repartidos en {num_sheets} hojas dentro de {stock_file}")

if __name__ == "__main__":
    procesar_stock()
