import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def comparar_stock():
    print(">>> Iniciando comparación de stock...")

    stock_file = "stock.xlsx"
    tienda_file = "tiendanube.csv"
    salida_file = "stock_diferencias.xlsx"

    # Leer stock.xlsx
    stock = pd.read_excel(stock_file)
    stock["default_code"] = stock["default_code"].astype(str).str.strip()
    stock["qty_available"] = stock["qty_available"].astype(float)

    # Leer tiendanube.csv
    tienda = pd.read_csv(tienda_file, sep=None, engine="python", encoding="latin1", on_bad_lines="skip")
    tienda["SKU"] = tienda["SKU"].astype(str).str.strip()
    tienda["Stock"] = tienda["Stock"].astype(float)

    # Comparación de stocks
    comparacion = stock.copy()
    comparacion["stock_tiendanube"] = comparacion["default_code"].map(dict(zip(tienda["SKU"], tienda["Stock"])))

    # Filtrar diferencias
    diferencias = comparacion[comparacion["qty_available"] != comparacion["stock_tiendanube"]]

    # Eliminar columnas innecesarias
    diferencias = diferencias.drop(columns=["id", "is_favorite"], errors="ignore")

    # Guardar en Excel
    with pd.ExcelWriter(salida_file, engine="openpyxl") as writer:
        diferencias.to_excel(writer, sheet_name="diferencias", index=False)

        # --- Productos faltantes ---
        skus_stock = set(stock["default_code"])
        skus_tienda = set(tienda["SKU"])

        faltan_en_tienda = skus_stock - skus_tienda
        faltan_en_stock = skus_tienda - skus_stock

        faltantes = pd.DataFrame(
            [(sku, "Falta en TiendaNube") for sku in faltan_en_tienda] +
            [(sku, "Falta en Odoo") for sku in faltan_en_stock],
            columns=["SKU", "Dónde falta"]
        )

        faltantes.to_excel(writer, sheet_name="faltantes", index=False)

    # Ajustar ancho de columnas
    wb = load_workbook(salida_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(salida_file)

    print(">>> Archivo generado:", salida_file)

if __name__ == "__main__":
    comparar_stock()
