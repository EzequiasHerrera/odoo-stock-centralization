import sys
import pandas as pd
from openpyxl import load_workbook

def generar_componentes_comb(boms_file):
    # Leer archivo BoMs (variable)
    wb = load_workbook(boms_file)
    ws_boms = wb.active  # Hoja 1 (BoMs vacías)
    
    # Leer archivo Odoo
    odoo = pd.read_excel("Odoo.xlsx")
    
    # Crear hoja de salida
    if "Componentes" in wb.sheetnames:
        ws_out = wb["Componentes"]
        wb.remove(ws_out)
    ws_out = wb.create_sheet("Componentes")
    
    # Encabezados
    headers = ["id","product_tmpl_id","product_tmpl_id/name","type",
               "bom_line_ids/product_id","bom_line_ids/product_qty","bom_line_ids/product_uom_id"]
    ws_out.append(headers)
    
    # Recorrer BoMs
    for row in ws_boms.iter_rows(min_row=2, values_only=True):
        bom_id, product_tmpl_id, product_name, bom_type = row[:4]
        
        if product_tmpl_id and "[" in product_tmpl_id and "]" in product_tmpl_id:
            # Extraer el SKU dentro de corchetes
            sku_full = product_tmpl_id.split("[")[1].split("]")[0].strip()
            
            # Eliminar primer segmento hasta el primer guión
            if "-" in sku_full:
                sku_real = sku_full.split("-", 1)[1]
            else:
                sku_real = sku_full
            
            # Buscar nombre en Odoo por SKU (columna F)
            odoo_skus = odoo.iloc[:,5].astype(str).str.strip()
            odoo_row = odoo.loc[odoo_skus == sku_real]
            if odoo_row.empty:
                print(f"[INFO] SKU '{sku_real}' no encontrado en Odoo.xlsx → se marca como DESCONOCIDO")
                prod_name = "DESCONOCIDO"
            else:
                prod_name = odoo_row.iloc[0,0]  # Columna A = nombre producto
            
            # Armar fila (cantidad como número 1)
            ws_out.append([bom_id, product_tmpl_id, product_name, bom_type,
                           f"[{sku_real}] {prod_name}", 1, "Unidades"])
    
    # Guardar archivo
    wb.save(boms_file)
    print(f"Procesamiento terminado. Resultados en hoja 'Componentes' de {boms_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python LdMComb.py <archivo_boms.xlsx>")
    else:
        generar_componentes_comb(sys.argv[1])
