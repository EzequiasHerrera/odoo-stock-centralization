import sys
import pandas as pd
from openpyxl import load_workbook

def generar_componentes(boms_file):
    # Leer archivo BoMs (variable)
    wb = load_workbook(boms_file)
    ws_boms = wb.active  # Hoja 1 (BoMs vacías)
    
    # Leer archivos auxiliares
    funsales = pd.read_excel("resultado_funsales.xlsx")
    odoo = pd.read_excel("Odoo.xlsx")
    
    # Crear hoja de salida
    if "Componentes" in wb.sheetnames:
        ws_out = wb["Componentes"]
        wb.remove(ws_out)
    ws_out = wb.create_sheet("Componentes")
    
    # Encabezados
    headers = ["id","product_tmpl_id","product_tmpl_id/name","type",
               "bom_line_ids/product_id","bom_line_ids/product_qty","bom_line_ids/product_uom_id"]
    ws_out.append(headers)
    
    # Recorrer BoMs
    for row in ws_boms.iter_rows(min_row=2, values_only=True):
        bom_id, product_tmpl_id, product_name, bom_type = row[:4]
        if product_tmpl_id and "[" in product_tmpl_id:
            ids_str = product_tmpl_id.split("[")[1].split("]")[0]
            ids = [id_val.strip() for id_val in ids_str.split("|") if id_val.strip() != ""]
            
            first = True
            for id_val in ids:
                # Buscar fila en resultado_funsales (columna A como string)
                funsales_ids = funsales.iloc[:,0].astype(str).str.strip()
                sku_row = funsales.loc[funsales_ids == id_val]
                if sku_row.empty:
                    print(f"ID {id_val} no encontrado en resultado_funsales")
                    continue
                
                # Tomar SKU: columna J (índice 9) o columna I (índice 8 si J está vacío)
                sku = sku_row.iloc[0,9] if pd.notna(sku_row.iloc[0,9]) and str(sku_row.iloc[0,9]).strip() != "" else sku_row.iloc[0,8]
                sku = str(sku).strip()
                
                # Buscar nombre en Odoo
                odoo_skus = odoo.iloc[:,5].astype(str).str.strip()
                odoo_row = odoo.loc[odoo_skus == sku]
                if odoo_row.empty:
                    print(f"SKU no encontrado en Odoo: {sku}")
                    prod_name = "DESCONOCIDO"
                else:
                    prod_name = odoo_row.iloc[0,0]  # Columna A = nombre producto
                
                # Armar fila (cantidad como número 1)
                if first:
                    ws_out.append([bom_id, product_tmpl_id, product_name, bom_type,
                                   f"[{sku}] {prod_name}", 1, "Unidades"])
                    first = False
                else:
                    ws_out.append(["","","","",
                                   f"[{sku}] {prod_name}", 1, "Unidades"])
    
    # Guardar archivo
    wb.save(boms_file)
    print(f"Procesamiento terminado. Resultados en hoja 'Componentes' de {boms_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python generar_boms.py <archivo_boms.xlsx>")
    else:
        generar_componentes(sys.argv[1])
